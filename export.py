import os
import pandas as pd
import shutil
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from dataverse_connector import NewDataverseConnector
from openpyxl.styles import Border, Side

def export_tournee_vers_excel(tournee_id=None, output_file=None):
    """
    Exporte les données d'une tournée spécifique vers un fichier Excel basé sur un modèle
    
    Args:
        tournee_id (str, optional): ID de la tournée à exporter (crcfe_idtournees)
        template_file (str): Chemin du fichier Excel modèle
        output_file (str, optional): Chemin du fichier Excel de sortie
    
    Returns:
        bool: True si l'exportation a réussi, False sinon
    """
    
    # Configuration du connecteur
    config = {
        'client_id': '8993b267-820e-4aef-851a-62158ddef76b',
        'tenant_id': 'df152455-73df-41a9-b48e-7fb075739495',
        'env_url': 'https://org51f7f291.crm4.dynamics.com/' 
    }
    
    # Créer et connecter le connecteur
    connector = NewDataverseConnector(**config)
    if not connector.connect():
        print("Échec de la connexion à Dataverse")
        return False
    
    # Construire le filtre pour récupérer la tournée spécifique
    filter_query = None
    if tournee_id:
        filter_query = f"crcfe_idtournees eq '{tournee_id}'"
    else:
        print("Veuillez spécifier l'ID de la tournée")
        return False
    
    # Récupérer les données de la tournée
    print(f"Récupération des données de la tournée avec filtre: {filter_query}")
    
    # 1. Récupérer les données de base de la tournée
    tournees_data = connector.get_table_data("crcfe_tournees", filter=filter_query, only_custom=True)
    if tournees_data is None or tournees_data.empty:
        print("Aucune tournée trouvée avec les critères spécifiés")
        return False
    
    print(f"Tournée trouvée: {len(tournees_data)} entrée(s)")
    
    if 'crcfe_tourneesid' not in tournees_data.columns:
        print("La colonne 'crcfe_tourneesid' n'est pas présente dans les données de tournée")
        return False
    
    tournee_unique_id = tournees_data['crcfe_tourneesid'].iloc[0]

    type_collecte = tournees_data['crcfe_type_collecte'].iloc[0] if 'crcfe_type_collecte' in tournees_data.columns else "OM"
    is_ep = type_collecte.upper() == "EP" if type_collecte else False

    print(f"Type de tournée détecté: {'EP' if is_ep else 'OM'}")
    template_file = "Suivi EP.xlsx" if is_ep else "Suivi OM.xlsx"
    
    if not os.path.exists(template_file):
        print(f"Le fichier modèle '{template_file}' n'existe pas")
        return False
    
    date_suivi = tournees_data['crcfe_date_suivi'].iloc[0] if 'crcfe_date_suivi' in tournees_data.columns else None
    heure_debut = tournees_data['crcfe_heure_debut'].iloc[0] if 'crcfe_heure_debut' in tournees_data.columns else None
    heure_fin = tournees_data['crcfe_heure_fin'].iloc[0] if 'crcfe_heure_fin' in tournees_data.columns else None
    nom_equipe = tournees_data['crcfe_nom_equipe'].iloc[0] if 'crcfe_nom_equipe' in tournees_data.columns else None
    immatriculation = tournees_data['crcfe_immatriculation_benne'].iloc[0] if 'crcfe_immatriculation_benne' in tournees_data.columns else None
    
    agents_filter = f"_crcfe_idtournees_value eq '{tournee_unique_id}'"
    agents_data = connector.get_table_data("crcfe_agentstournees", filter=agents_filter, only_custom=False)
    if agents_data is None:
        print("Erreur lors de la récupération des agents de la tournée")
        agents_data = pd.DataFrame()
    else:
        print(f"Agents trouvés: {len(agents_data)} entrée(s)")
    
    agents_noms = []
    if not agents_data.empty:
        agent_ids = []
        if '_crcfe_id_agent_value' in agents_data.columns:
            agent_ids = agents_data['_crcfe_id_agent_value'].dropna().tolist()
        elif 'crcfe_id_agent' in agents_data.columns:
            agent_ids = agents_data['crcfe_id_agent'].dropna().tolist()
        
        if agent_ids:
            # Construire un filtre pour récupérer les agents
            agent_filter = " or ".join([f"new_agentsid eq '{a_id}'" for a_id in agent_ids])
            agents_info = connector.get_table_data("new_agents", filter=agent_filter, only_custom=False)
            
            if agents_info is not None and not agents_info.empty:
                
                # Extraire les noms et prénoms
                for _, agent in agents_info.iterrows():
                    nom = ""
                    prenom = ""
                    
                    # Rechercher les colonnes de nom et prénom
                    for col in agent.index:
                        if 'nom' in col.lower() and not 'prenom' in col.lower():
                            nom = agent[col] if pd.notna(agent[col]) else ""
                        elif 'prenom' in col.lower() or 'prénom' in col.lower():
                            prenom = agent[col] if pd.notna(agent[col]) else ""
                    
                    # Ajouter le nom complet
                    if prenom and nom:
                        agents_noms.append(f"{prenom} {nom}")
                    elif prenom:
                        agents_noms.append(prenom)
                    elif nom:
                        agents_noms.append(nom)

    agents_str = ", ".join(agents_noms) if agents_noms else ""
    print(f"Noms des agents: {agents_str}")
    
    # 3. Récupérer les bacs associés à cette tournée
    bacs_filter = f"_crcfe_id_tournee_value eq '{tournee_unique_id}'"
    bacs_data = connector.get_table_data("new_bacs", filter=bacs_filter, only_custom=False)
    
    if bacs_data is None:
        print("Erreur lors de la récupération des bacs de la tournée")
        bacs_data = pd.DataFrame()
    else:
        print(f"Bacs trouvés: {len(bacs_data)} entrée(s)")
    
    # 4. Récupérer les adresses associées aux bacs
    adresses_data = None
    bacs_adresses = []
    
    # Si des bacs ont été trouvés, récupérer leurs adresses
    if not bacs_data.empty:
        # Récupérer les IDs des bacs
        bac_ids = []
        bac_ids.extend(bacs_data['new_bacsid'].dropna().tolist())
        
        print(f"IDs des bacs trouvés: {bac_ids}")
        # Récupérer les adresses directement à partir des bacs
        adresse_ids = []
        
        # Chercher le champ de lookup pour les adresses
        adresse_ids.extend(bacs_data["_crcfe_adressebac_value"].dropna().tolist())
        
        if adresse_ids:
            # Construire un filtre pour récupérer les adresses
            adresse_filter = " or ".join([f"crcfe_listeadressesbacsid eq '{a_id}'" for a_id in adresse_ids])
            adresses_data = connector.get_table_data("crcfe_listeadressesbacs", filter=adresse_filter, only_custom=False)
            
            if adresses_data is not None and not adresses_data.empty:
                print(f"Adresses trouvées directement à partir des bacs: {len(adresses_data)} entrée(s)")
                
            # Pour chaque bac, trouver son adresse associée et générer l'entrée pour Excel
            for i, bac_row in bacs_data.iterrows():
                adresse_id = bac_row.get('_crcfe_adressebac_value')
                if adresse_id:
                    # Trouver l'adresse correspondante
                    adresse = adresses_data[adresses_data['crcfe_listeadressesbacsid'] == adresse_id]
                    
                    if not adresse.empty:
                        commune = adresse['crcfe_commune'].iloc[0] if 'crcfe_commune' in adresse.columns else ""
                        numero = adresse['crcfe_numerorue'].iloc[0] if 'crcfe_numerorue' in adresse.columns else ""
                        bis_ter = adresse['crcfe_bister'].iloc[0] if 'crcfe_bister' in adresse.columns else ""
                        nom_rue = adresse['crcfe_nomrue'].iloc[0] if 'crcfe_nomrue' in adresse.columns else ""
                        type_habitat = adresse['crcfe_typehabitat'].iloc[0] if 'crcfe_typehabitat' in adresse.columns else ""
                        
                        # Récupérer le volume du bac et le taux de remplissage
                        bac_info = {
                            'COMMUNE': commune,
                            'N°': numero,
                            'BIS_TER': bis_ter,
                            'NOM_RUE': nom_rue,
                            'TYPE_HABITAT': type_habitat,
                        }
                        
                        # Ajouter les champs spécifiques selon le type de tournée
                        if is_ep:
                            # Récupérer l'action EP si disponible
                            action_ep = ""
                            for col in bac_row.index:
                                if 'action_ep' in col.lower():
                                    action_ep = bac_row[col] if pd.notna(bac_row[col]) else ""
                            
                            # Étendre avec les champs EP
                            bac_info.update({
                                'ACTIONS': action_ep,
                                'SACS_OM': "",
                                'DEEE': "",
                                'DECHETS_TOXIQUES': "",
                                'GRAVATS': "",
                                'DECHETS_VEGETAUX': "",
                                'VERRE': "",
                                'CARTON_MOUILLE': "",
                                'VETEMENT': "",
                                'AUTRES': "",
                                '*DECHETS': "",
                                'OBSERVATION': "",
                                '*ENSEIGNE': "",
                                'Commentaires': ""
                            })
                        else:
                            # Pour OM, ajouter les informations de volume et taux
                            volume_bac = ""
                            taux_remplissage = ""
                            commentaire = ""
                            
                            for col in bac_row.index:
                                if 'volume' in col.lower():
                                    volume_bac = bac_row[col] if pd.notna(bac_row[col]) else ""
                                elif 'taux' in col.lower() or 'remplissage' in col.lower():
                                    taux_remplissage = bac_row[col] if pd.notna(bac_row[col]) else ""
                                elif 'commentaire' in col.lower():
                                    commentaire = bac_row[col] if pd.notna(bac_row[col]) else ""
                            
                            bac_info.update({
                                'Volume du Bac': volume_bac,
                                'TAUX': taux_remplissage,
                                'Commentaires': commentaire
                            })
                        
                        # Ajouter cette ligne à notre liste d'adresses de bacs
                        bacs_adresses.append(bac_info)
    
    # Créer un DataFrame pour les bacs avec leurs adresses
    bacs_df = pd.DataFrame(bacs_adresses)

    vidage_filter = f"_crcfe_idtournees_value eq '{tournee_unique_id}'"
    vidage_data = connector.get_table_data("new_vidages", filter=vidage_filter, only_custom=False)

    if vidage_data is None:
        print("Erreur lors de la récupération des vidages de la tournée")
        heures_vidage = ""
    elif vidage_data.empty:
        print("Aucun vidage trouvé pour cette tournée")
        heures_vidage = ""
    else:
        print(f"Vidages trouvés: {len(vidage_data)} entrée(s)")

        # Récupérer les heures de vidage
        heures_list = []
        for _, vidage in vidage_data.iterrows():
            heure_vidage = vidage['new_heure_vidage']
            if heure_vidage:
                try:
                    if isinstance(heure_vidage, str) and 'T' in heure_vidage:
                        dt_heure_vidage = datetime.fromisoformat(heure_vidage.replace('Z', '+00:00'))
                        dt_heure_vidage = dt_heure_vidage + timedelta(hours=2)
                        heure_vidage = dt_heure_vidage.strftime('%H:%M')
                    heures_list.append(heure_vidage)
                except Exception as e:
                    print(f"Erreur lors de la conversion de l'heure de vidage: {e}")
                    heures_list.append(str(heure_vidage))
    
        heures_vidage = ", ".join(heures_list) if heures_list else ""
    
    print(bacs_df)

    # Définir le nom du fichier de sortie s'il n'est pas spécifié
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        type_str = "EP" if is_ep else "OM"
        tournee_identifier = tournee_id if tournee_id else "inconnue"
        output_file = f"Suivi_{type_str}_Tournee_{tournee_identifier}_{timestamp}.xlsx"

    # Copier le fichier template
    shutil.copy2(template_file, output_file)
    print(f"Création du fichier de sortie: {output_file}")

    # Ouvrir le fichier copié avec openpyxl
    wb = load_workbook(output_file)
    sheet_name = "Suivi de collecte EP" if is_ep else "Suivi de collecte OM"
    sheet = wb[sheet_name]

    if date_suivi:
        try:
            if isinstance(date_suivi, str) and 'T' in date_suivi:
                dt_date_suivi = datetime.fromisoformat(date_suivi.replace('Z', '+00:00'))
                dt_date_suivi = dt_date_suivi + timedelta(hours=2)
                date_suivi = dt_date_suivi.strftime('%d/%m/%Y')
            sheet['C2'] = date_suivi
        except Exception as e:
            print(f"Erreur lors de la conversion de la date: {e}")
            sheet['C2'] = str(date_suivi)

    equipe_agents = f"{nom_equipe or ''}"
    if agents_str:
        equipe_agents += f" / {agents_str}"
    sheet['C3'] = equipe_agents

    if immatriculation:
        sheet['C4'] = immatriculation

    if heure_debut:
        try:
            if isinstance(heure_debut, str) and 'T' in heure_debut:
                dt_heure_debut = datetime.fromisoformat(heure_debut.replace('Z', '+00:00'))
                dt_heure_debut = dt_heure_debut + timedelta(hours=2)
                heure_debut = dt_heure_debut.strftime('%H:%M')
            sheet['C5'] = heure_debut
        except Exception as e:
            print(f"Erreur lors de la conversion de l'heure de début: {e}")
            sheet['C5'] = str(heure_debut)

    if heure_fin:
        try:
            if isinstance(heure_fin, str) and 'T' in heure_fin:
                dt_heure_fin = datetime.fromisoformat(heure_fin.replace('Z', '+00:00'))
                dt_heure_fin = dt_heure_fin + timedelta(hours=2)
                heure_fin = dt_heure_fin.strftime('%H:%M')
            sheet['C6'] = heure_fin
        except Exception as e:
            print(f"Erreur lors de la conversion de l'heure de fin: {e}")
            sheet['C6'] = str(heure_fin)
    
    thick_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Dans la partie où vous ajoutez les heures de vidage
    if heures_list:
        # Mettre chaque heure de vidage dans une cellule séparée
        for idx, heure in enumerate(heures_list):
            # Commencer à la colonne C (index 3) et continuer vers la droite
            col_idx = 3 + idx  # C=3, D=4, E=5, etc.
            col_letter = chr(65 + col_idx - 1)  # Convertir l'index en lettre (A=65 en ASCII)
            cell = sheet[f'{col_letter}7']
            cell.value = heure
            
            # Appliquer les bordures en gras à la cellule
            cell.border = thick_border

    # Remplir les données des bacs
    if not bacs_df.empty:
        # Commencer à remplir à partir de la ligne 11
        start_row = 11
        
        print(f"Remplissage des données de {len(bacs_df)} bacs dans le fichier Excel...")
        
        # Définir le mapping des colonnes selon le type de tournée
        if is_ep:
            column_mapping = {
                'COMMUNE': 1,          
                'N°': 2,               
                'BIS_TER': 3,          
                'NOM_RUE': 4,          
                'TYPE_HABITAT': 5,     
                'ACTIONS': 6,          
                'SACS_OM': 7,          
                'DEEE': 8,             
                'DECHETS_TOXIQUES': 9, 
                'GRAVATS': 10,         
                'DECHETS_VEGETAUX': 11,
                'VERRE': 12,           
                'CARTON_MOUILLE': 13,  
                'VETEMENT': 14,        
                'AUTRES_DECHETS': 15,  
                'OBSERVATION_ENSEIGNE': 16,   
                'Commentaires': 17
            }
        else:
            column_mapping = {
                'COMMUNE': 1,
                'N°': 2,
                'BIS_TER': 3,
                'NOM_RUE': 4,
                'TYPE_HABITAT': 5,
                'Volume du Bac': 6,
                'TAUX': 7,
                'Commentaires': 8
            }
        
        for i, row in bacs_df.iterrows():
            for col_name, col_index in column_mapping.items():
                sheet.cell(row=start_row + i, column=col_index, value=row.get(col_name, ''))

    wb.save(output_file)
    print(f"Le fichier Excel '{output_file}' a été créé avec succès.")

    return True

def main():
    """Fonction principale pour interagir avec l'utilisateur"""
    print("=== Exportation des données d'une tournée spécifique vers Excel ===")

    tournee_id = input("Entrez l'ID de la tournée à exporter: ")
    success = export_tournee_vers_excel(tournee_id=tournee_id)
    
    if success:
        print("Exportation terminée avec succès!")
    else:
        print("L'exportation a échoué. Veuillez vérifier les messages d'erreur ci-dessus.")

if __name__ == "__main__":
    main()